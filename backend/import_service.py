"""
资产批量导入服务模块

负责：
1. Excel 文件解析（pandas）+ 模糊列头匹配
2. 品类自动推断（基于文件名关键词）
3. 数据验证（AssetCreate Pydantic + 自定义规则）
4. 未知列数据捕获到 additional_info JSON 字段
5. 批量写入（事务性）+ 分级校验（冲突跳过并记录）
6. 导入模板生成（openpyxl）
"""

import re
import io
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pydantic import ValidationError
from sqlalchemy.orm import Session
from typing import Optional

import models
import schemas


# ========== 精确列头映射表（中文 → 英文字段名）==========
# 这是权威映射，模糊匹配以此为基础扩展
COLUMN_MAPPING: dict[str, str] = {
    "资产编号": "asset_tag",
    "品类": "category",
    "工号": "employee_id",
    "姓名": "employee_name",
    "部门": "department",
    "直属领导": "supervisor",
    "资产名": "hostname",
    "状态": "status",
    "型号": "model",
    "序列号": "serial_number",
    "品牌": "brand",
    "MAC地址": "mac_address",
    "IP地址": "ip_address",
    "备注": "notes",
    "系统版本": "system_version",
    "杀毒软件": "antivirus_software",
    "锁号": "lock_number",
    "位置": "location",
    "数量": "quantity",
    "采购日期": "purchase_date",
    "固定资产编号": "fixed_asset_number",
    "PO号": "po_number",
}

# ========== 模糊列头别名表（常见变体 → 标准中文列头）==========
# 用于识别用户自定义表头中的常见写法差异
FUZZY_ALIAS: dict[str, str] = {
    # 资产编号
    "资产号": "资产编号",
    "编号": "资产编号",
    "资产标签": "资产编号",
    "asset_tag": "资产编号",
    "资产tag": "资产编号",
    # 品类
    "类别": "品类",
    "分类": "品类",
    "设备类型": "品类",
    "类型": "品类",
    "category": "品类",
    # 品牌
    "厂商": "品牌",
    "厂家": "品牌",
    "brand": "品牌",
    # 型号
    "规格": "型号",
    "产品型号": "型号",
    "model": "型号",
    "计算机型号": "型号",
    "机型": "型号",
    # 序列号
    "SN": "序列号",
    "sn": "序列号",
    "S/N": "序列号",
    "serial": "序列号",
    "序号": "序列号",
    "机器序列号": "序列号",
    # 状态
    "使用状态": "状态",
    "资产状态": "状态",
    "status": "状态",
    # 使用人（标准模板列头为“姓名”，旧列头继续作为别名）
    "使用人": "姓名",
    "员工姓名": "姓名",
    "姓名": "姓名",
    "用户": "姓名",
    "领用人": "姓名",
    "持有人": "姓名",
    "登记姓名": "姓名",
    #直属领导
    "部门主管": "直属领导",
    # 工号
    "员工工号": "工号",
    "员工编号": "工号",
    "人员编号": "工号",
    # 部门
    "所属部门": "部门",
    "归属部门": "部门",
    "department": "部门",
    # 资产名/主机名
    "主机名": "资产名",
    "设备名": "资产名",
    "设备名称": "资产名",
    "hostname": "资产名",
    "计算机名": "资产名",
    # MAC地址
    "mac": "MAC地址",
    "MAC": "MAC地址",
    "网卡地址": "MAC地址",
    "物理地址": "MAC地址",
    "Wireless mac": "MAC地址",
    # IP地址
    "IP": "IP地址",
    "ip": "IP地址",
    "内网IP": "IP地址",
    # 固定资产编号
    "固资编号": "固定资产编号",
    "固定资产号": "固定资产编号",
    "财务编号": "固定资产编号",
    "Fixed assets": "资产编号",
    # 备注
    "说明": "备注",
    "描述": "备注",
    "remark": "备注",
    "notes": "备注",
    # 系统版本
    "操作系统": "系统版本",
    "OS": "系统版本",
    "os版本": "系统版本",
    "System": "系统版本",
    # 位置
    "存放位置": "位置",
    "资产位置": "位置",
    "放置位置": "位置",
    "使用位置": "位置",
    "FA使用区域": "位置",
    # 采购日期
    "购买日期": "采购日期",
    "入库日期": "采购日期",
    "购置日期": "采购日期",
    # PO号
    "po号": "PO号",
    "PO": "PO号",
    "po": "PO号",
    "采购订单号": "PO号",
    "采购单号": "PO号",
    "订单号": "PO号",
}

# 品类名称标准化映射（Excel 中的变体 → 系统标准值）
CATEGORY_NORMALIZE: dict[str, str] = {
    "台式电脑": "台式机",
    "台式": "台式机",
    "desktop": "台式机",
    "笔记本": "笔记本电脑",
    "laptop": "笔记本电脑",
    "notebook": "笔记本电脑",
    "平板": "移动设备",
    "ipad": "移动设备",
    "tablet": "移动设备",
    "PAD": "移动设备",
}

# n/a 类占位符，导入时统一转为 None
NA_VALUES: set[str] = {"n/a", "na", "none", "null", "-", "无", "暂无", "不适用"}

# 反向映射：英文字段名 → 中文列头（用于错误消息）
REVERSE_MAPPING: dict[str, str] = {v: k for k, v in COLUMN_MAPPING.items()}
# 注意：品类在启用文件名推断时可以不强制要求列头存在
REQUIRED_COLUMNS: set[str] = {"资产编号", "状态"}

# 合法状态值集合
VALID_STATUSES: set[str] = {"闲置", "使用中", "维修中", "报废"}

# asset_tag 格式正则：ZS-XXXX-NNNNNN（如 ZS-PC26-000001）
ASSET_TAG_PATTERN: re.Pattern = re.compile(r"^ZS-[A-Za-z0-9]{4}-\d{6}$")

# ========== 品类自动推断规则（文件名关键词 → 品类）==========
# 按优先级排列，先匹配先生效
FILENAME_CATEGORY_RULES: list[tuple[list[str], str]] = [
    (["笔记本", "laptop", "notebook", "nb"],        "笔记本电脑"),
    (["台式", "台式电脑", "desktop", "pc"],              "台式机"),
    (["服务器", "server", "srv"],                    "服务器"),
    (["移动设备", "pad", "ipad", "平板", "tablet"],  "移动设备"),
    (["手机", "phone", "mobile"],                    "手机"),
    (["显示器", "monitor", "display"],               "显示器"),
    (["打印机", "printer"],                          "打印机"),
    (["网络", "network", "路由", "交换机", "switch"], "网络设备"),
    (["鼠标", "mouse", "键盘", "keyboard"],          "无线鼠标"),
]

# 模板示例数据
TEMPLATE_EXAMPLE_ROW = {
    "资产编号": "ZS-NB26-000001",
    "品类": "笔记本电脑",
    "工号": "",
    "姓名": "",
    "部门": "",
    "直属领导": "",
    "资产名": "LAPTOP-001",
    "状态": "闲置",
    "型号": "X1 Carbon",
    "序列号": "SN123456789",
    "品牌": "ThinkPad",
    "MAC地址": "AA:BB:CC:DD:EE:FF",
    "IP地址": "192.168.1.100",
    "备注": "示例数据，请删除后填写实际数据",
    "系统版本": "",
    "杀毒软件": "",
    "锁号": "",
    "位置": "",
    "数量": "",
    "采购日期": "",
    "固定资产编号": "FA-2024-0001",
    "PO号": "12000327",
}


def infer_category_from_filename(filename: str) -> Optional[str]:
    """
    根据文件名关键词推断资产品类。

    匹配规则不区分大小写，按 FILENAME_CATEGORY_RULES 顺序优先匹配。

    参数:
        filename: 上传的文件名（含扩展名）

    返回:
        推断出的品类字符串，无匹配时返回 None
    """
    name_lower = filename.lower()
    for keywords, category in FILENAME_CATEGORY_RULES:
        if any(kw.lower() in name_lower for kw in keywords):
            return category
    return None


def resolve_header(raw_header: str) -> Optional[str]:
    """
    将原始列头（可能是别名或变体）解析为标准中文列头，再映射到英文字段名。

    解析顺序：
    1. 精确匹配 COLUMN_MAPPING（直接返回英文字段名）
    2. 精确匹配 FUZZY_ALIAS → 再查 COLUMN_MAPPING
    3. 去除空格后重试上述两步
    4. 均未命中 → 返回 None（该列视为未知列，存入 additional_info）

    参数:
        raw_header: Excel 中的原始列头字符串

    返回:
        英文字段名，或 None（未知列）
    """
    h = str(raw_header).strip()

    # 1. 精确匹配标准列头
    if h in COLUMN_MAPPING:
        return COLUMN_MAPPING[h]

    # 2. 精确匹配别名
    if h in FUZZY_ALIAS:
        std = FUZZY_ALIAS[h]
        return COLUMN_MAPPING.get(std)

    # 3. 去除所有空格后重试
    h_nospace = h.replace(" ", "").replace("\u3000", "")
    if h_nospace in COLUMN_MAPPING:
        return COLUMN_MAPPING[h_nospace]
    if h_nospace in FUZZY_ALIAS:
        std = FUZZY_ALIAS[h_nospace]
        return COLUMN_MAPPING.get(std)

    # 4. 未知列
    return None


def parse_excel(
    file_content: bytes,
    filename: str = ""
) -> tuple[list[str], list[dict[str, Optional[str]]], Optional[str]]:
    """
    解析 Excel 文件，支持模糊列头匹配和未知列捕获。

    流程:
    1. pandas 读取，所有列强制为字符串
    2. 检查必填列头（资产编号、状态）
    3. 尝试从文件名推断品类（inferred_category）
    4. 对每个列头调用 resolve_header：
       - 已知列 → 映射为英文字段名
       - 未知列 → 保留原始列头，后续存入 additional_info
    5. 构建行数据字典，未知列数据聚合到 "_extra" 键

    参数:
        file_content: 文件二进制内容
        filename: 原始文件名（用于品类推断）

    返回:
        (mapped_headers, rows, inferred_category)
        - mapped_headers: 已知列的英文字段名列表
        - rows: 行数据字典列表，含 "_extra" 键（未知列数据）
        - inferred_category: 从文件名推断的品类，无法推断时为 None

    异常:
        ValueError: 缺少必填列头或文件无法解析
    """
    try:
        df = pd.read_excel(io.BytesIO(file_content), dtype=str)
    except Exception as e:
        raise ValueError(f"文件解析失败: {str(e)}")

    original_headers = list(df.columns)

    # ── 检查必填列头（模糊匹配后检查）──
    # 将原始列头解析为标准中文名，用于必填检查
    resolved_std_names: set[str] = set()
    for h in original_headers:
        h_stripped = str(h).strip()
        # 先看是否直接在标准列头里
        if h_stripped in COLUMN_MAPPING:
            resolved_std_names.add(h_stripped)
        elif h_stripped in FUZZY_ALIAS:
            resolved_std_names.add(FUZZY_ALIAS[h_stripped])

    missing_required = REQUIRED_COLUMNS - resolved_std_names
    if missing_required:
        raise ValueError(f"缺少必填列头: {', '.join(sorted(missing_required))}（支持别名，如'编号'可代替'资产编号'）")

    # ── 从文件名推断品类 ──
    inferred_category = infer_category_from_filename(filename) if filename else None

    # ── 构建列头解析结果 ──
    # known_map: 原始列头 → 英文字段名（已知列）
    # unknown_headers: 未知列的原始列头列表
    known_map: dict[str, str] = {}
    unknown_headers: list[str] = []

    for h in original_headers:
        en_field = resolve_header(str(h))
        if en_field is not None:
            # 同一英文字段可能被多个别名命中，取第一个
            if en_field not in known_map.values():
                known_map[str(h)] = en_field
            else:
                # 重复映射的列视为未知列，避免覆盖
                unknown_headers.append(str(h))
        else:
            unknown_headers.append(str(h))

    mapped_headers = list(known_map.values())

    # ── 构建行数据 ──
    rows = []
    for _, row in df.iterrows():
        row_dict: dict[str, Optional[str]] = {}

        # 已知列
        for orig_h, en_field in known_map.items():
            val = row[orig_h]
            if (not isinstance(val, str) and pd.isna(val)) or (isinstance(val, str) and val.strip() == ""):
                row_dict[en_field] = None
            else:
                row_dict[en_field] = str(val).strip()

        # 未知列 → 聚合到 _extra
        extra: dict[str, str] = {}
        for orig_h in unknown_headers:
            val = row[orig_h]
            if (not isinstance(val, str) and pd.isna(val)) or (isinstance(val, str) and val.strip() == ""):
                continue  # 空值不存入 extra
            extra[str(orig_h)] = str(val).strip()
        if extra:
            row_dict["_extra"] = extra  # type: ignore

        rows.append(row_dict)

    return mapped_headers, rows, inferred_category


def validate_rows(
    rows: list[dict[str, Optional[str]]],
    db: Session,
    inferred_category: Optional[str] = None,
) -> tuple[list[dict], list[dict]]:
    """
    逐行验证数据，支持品类推断填充和分级校验（冲突跳过并记录）。

    验证流程（每行按顺序）:
    1. 构建 row_data，空字符串转 None
    2. 品类推断：若行内无 category 且 inferred_category 不为 None，自动填充
    3. 提取 _extra 未知列数据，准备存入 additional_info
    4. Pydantic 验证（AssetCreate）
    5. 自定义分级校验：
       a. asset_tag 格式验证
       b. asset_tag 文件内去重
       c. asset_tag 数据库唯一性（冲突 → 跳过并记录，不中断整批）
       d. status 合法性
       e. serial_number 数据库唯一性（冲突 → 跳过并记录）
       f. 使用中时 employee_name 必填

    参数:
        rows: parse_excel 返回的行数据列表
        db: 数据库 Session
        inferred_category: 从文件名推断的品类（可为 None）

    返回:
        (valid_rows, errors)
    """
    valid_rows: list[dict] = []
    errors: list[dict] = []

    seen_asset_tags: dict[str, int] = {}

    for idx, row in enumerate(rows):
        row_number = idx + 2  # Excel 行号从 2 开始
        asset_tag_val = row.get("asset_tag")

        def add_error(message: str, skip_reason: str = ""):
            """记录失败行，skip_reason 用于区分冲突跳过和格式错误"""
            errors.append({
                "row_number": row_number,
                "asset_tag": asset_tag_val,
                "message": message,
                "skip_reason": skip_reason,  # "conflict" | "format" | "validation" | ""
            })

        # ── 步骤 1：构建 row_data，空字符串转 None ──
        extra_data: dict = row.pop("_extra", {}) if "_extra" in row else {}  # type: ignore
        row_data: dict = {}
        for k, v in row.items():
            if v is None or (isinstance(v, str) and v.strip() == ""):
                row_data[k] = None
            else:
                # 清理 n/a 类占位符
                v_stripped = str(v).strip()
                if v_stripped.lower() in NA_VALUES:
                    row_data[k] = None
                else:
                    row_data[k] = v_stripped

        # ── 步骤 2：品类标准化 ──
        category_raw = row_data.get("category")
        if category_raw:
            category_normalized = CATEGORY_NORMALIZE.get(category_raw, category_raw)
            row_data["category"] = category_normalized

        # ── 步骤 3：品类推断填充 ──
        if not row_data.get("category") and inferred_category:
            row_data["category"] = inferred_category

        # ── 步骤 4：Pydantic 验证（AssetCreate）──
        # additional_info 不在 AssetCreate 的校验范围内，先剔除
        pydantic_data = {k: v for k, v in row_data.items() if k != "additional_info"}
        try:
            schemas.AssetCreate(**pydantic_data)
        except ValidationError as e:
            messages = []
            for err in e.errors():
                msg = err.get("msg", "")
                if msg.startswith("Value error, "):
                    msg = msg[len("Value error, "):]
                messages.append(msg)
            add_error("; ".join(messages), skip_reason="validation")
            continue

        # ── 步骤 5a：asset_tag 格式验证 ──
        if not asset_tag_val or not ASSET_TAG_PATTERN.match(str(asset_tag_val)):
            add_error("资产编号格式错误，应为 ZS-XXXX-NNNNNN（如 ZS-PC26-000001）", skip_reason="format")
            continue

        # ── 步骤 5b：asset_tag 文件内去重 ──
        if asset_tag_val in seen_asset_tags:
            first_row = seen_asset_tags[asset_tag_val]
            add_error(
                f"资产编号 {asset_tag_val} 在文件中重复（首次出现于第 {first_row} 行）",
                skip_reason="conflict",
            )
            continue
        seen_asset_tags[asset_tag_val] = row_number

        # ── 步骤 5c：asset_tag 数据库唯一性（冲突跳过，不中断整批）──
        existing_asset = db.query(models.Asset).filter(
            models.Asset.asset_tag == asset_tag_val,
            models.Asset.is_deleted == False,
        ).first()
        if existing_asset:
            add_error(
                f"资产编号 {asset_tag_val} 已存在于数据库，已跳过该行",
                skip_reason="conflict",
            )
            continue

        # ── 步骤 5d：status 合法性 ──
        status_val = row_data.get("status")
        if status_val:
            status_stripped = status_val.strip()
            if status_stripped not in VALID_STATUSES:
                add_error(
                    f"非法状态值「{status_val}」，允许: 闲置, 使用中, 维修中, 报废",
                    skip_reason="validation",
                )
                continue
            row_data["status"] = status_stripped

        # ── 步骤 5e：serial_number 数据库唯一性（冲突跳过）──
        serial_number_val = row_data.get("serial_number")
        if serial_number_val:
            existing_sn = db.query(models.Asset).filter(
                models.Asset.serial_number == serial_number_val,
                models.Asset.is_deleted == False,
            ).first()
            if existing_sn:
                add_error(
                    f"序列号 {serial_number_val} 已存在于数据库，已跳过该行",
                    skip_reason="conflict",
                )
                continue

        # ── 步骤 5f：使用中时 employee_name 必填 ──
        if row_data.get("status") == "使用中":
            emp_name = row_data.get("employee_name")
            if not emp_name or not str(emp_name).strip():
                add_error("状态为「使用中」时，使用人为必填项", skip_reason="validation")
                continue

        # ── 步骤 6：将未知列数据合并到 additional_info ──
        if extra_data:
            existing_info = row_data.get("additional_info") or {}
            row_data["additional_info"] = {**existing_info, **extra_data}

        # 通过所有验证
        valid_rows.append({**row_data, "_row_number": row_number})

    return valid_rows, errors


def _sync_warehouse_quantity_import(db: Session, category: str, delta: int, operator_name: str = None):
    """
    同步库房资产可用数量（导入服务内部使用，不单独 commit）。
    与 main.py 中的 _sync_warehouse_quantity 逻辑一致，但独立实现避免循环导入。
    """
    if delta == 0:
        return

    # 资产品类 → 库房品类 映射（与 main.py 保持一致）
    ASSET_TO_WAREHOUSE_CATEGORY = {
        "台式机":    "计算机设备",
        "笔记本电脑": "计算机设备",
        "显示器":    "显示设备",
        "移动设备":  "移动设备",
        "手机":      "移动设备",
        "无线鼠标":  "输入设备",
        "打印机":    "其他配件",
        "网络设备":  "网络设备",
    }
    warehouse_category = ASSET_TO_WAREHOUSE_CATEGORY.get(category, category)

    wh_asset = (
        db.query(models.WarehouseAsset)
        .filter(models.WarehouseAsset.category == warehouse_category)
        .order_by(models.WarehouseAsset.available_quantity.desc())
        .with_for_update()
        .first()
    )
    if not wh_asset:
        return
    old_available = wh_asset.available_quantity
    old_allocated = wh_asset.allocated_quantity
    new_available = max(0, wh_asset.available_quantity + delta)
    wh_asset.available_quantity = new_available
    wh_asset.allocated_quantity = max(0, wh_asset.total_quantity - wh_asset.available_quantity)
    db.flush()
    desc = (
        f"批量导入联动更新: 可用数量 {old_available} → {wh_asset.available_quantity}, "
        f"已分配 {old_allocated} → {wh_asset.allocated_quantity}"
    )
    wh_log = models.WarehouseAssetLog(
        asset_id=wh_asset.id,
        action="入库（批量导入）",
        description=desc,
        operator=operator_name,
    )
    db.add(wh_log)
    db.flush()


def bulk_insert_assets(
    valid_rows: list[dict],
    db: Session,
    current_user: models.User,
) -> int:
    """
    在单个事务中批量写入资产记录，并创建审计日志。

    对每条记录:
    1. 移除内部字段（_row_number）
    2. 空字符串的唯一字段（serial_number, fixed_asset_number）转为 None
    3. additional_info 已是 dict，直接写入 JSON 列
    4. 创建 Asset 实例并 add 到 session
    5. 创建 AssetLog（action="批量导入"）
    6. 如果 status 为"闲置"，同步库房数量

    参数:
        valid_rows: 通过验证的行数据（含 _row_number 内部字段）
        db: 数据库 Session（调用方负责 commit/rollback）
        current_user: 当前操作用户

    返回:
        成功写入的记录数
    """
    operator_name = current_user.full_name or current_user.username
    inserted_count = 0

    for row in valid_rows:
        asset_data = {k: v for k, v in row.items() if not k.startswith("_")}

        # 空字符串的唯一字段转为 None，避免唯一约束冲突
        for field in ["serial_number", "fixed_asset_number"]:
            if field in asset_data and not asset_data[field]:
                asset_data[field] = None

        # additional_info 已是 dict（或 None），SQLAlchemy JSON 列直接接受
        db_asset = models.Asset(**asset_data)
        db.add(db_asset)
        db.flush()

        # 记录导入日志，附带 additional_info 摘要
        extra_note = ""
        if asset_data.get("additional_info"):
            keys = list(asset_data["additional_info"].keys())
            extra_note = f"，附加字段: {', '.join(keys)}"
        log = models.AssetLog(
            asset_id=db_asset.id,
            action="批量导入",
            description=f"通过 Excel 导入创建资产 {db_asset.asset_tag}{extra_note}",
            operator=operator_name,
        )
        db.add(log)
        db.flush()

        if db_asset.status == "闲置":
            _sync_warehouse_quantity_import(db, db_asset.category, +1, operator_name)

        inserted_count += 1

    # 整次导入的操作日志
    op_log = models.OperationLog(
        user_id=current_user.id,
        action="import",
        resource_type="asset",
        description=f"批量导入资产，成功写入 {inserted_count} 条",
    )
    db.add(op_log)
    db.flush()

    return inserted_count


def generate_template() -> bytes:
    """
    使用 openpyxl 生成导入模板 .xlsx 文件的二进制内容。

    模板包含:
    - 第 1 行: 中文列头（标准列头），加粗高亮
    - 第 2 行: 示例数据

    返回:
        .xlsx 文件的 bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "资产导入模板"

    headers = list(COLUMN_MAPPING.keys())

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx, value=TEMPLATE_EXAMPLE_ROW.get(header, ""))

    for col_idx, header in enumerate(headers, start=1):
        example_val = str(TEMPLATE_EXAMPLE_ROW.get(header, ""))
        col_width = max(len(header) * 2, len(example_val) + 2, 12)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
