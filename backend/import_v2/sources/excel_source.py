"""
import_v2.sources.excel_source
================================
ExcelSource：从 .xlsx 文件读取数据，输出 AssetRecord 列表。

职责（单一）：
  - 读取 Excel 二进制内容（pandas）
  - 模糊列头匹配：将中文/英文/别名列头解析为标准英文字段名
  - 必填列头检查（资产编号、状态）
  - 从文件名推断品类（inferred_category）
  - 构建 AssetRecord：填充 raw_fields、fields、extra_fields
  - 生成导入模板 .xlsx（openpyxl）

不做：
  - 不做数据标准化（那是 Normalizer 的职责）
  - 不访问数据库
  - 不做业务校验
"""

from __future__ import annotations

import io
from typing import Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ..domain_models import AssetRecord


# ===========================================================================
# 列头映射表（从 import_service.py 迁移，保持原有覆盖范围）
# ===========================================================================

# 精确列头映射表（中文标准列头 → 英文字段名）
COLUMN_MAPPING: dict[str, str] = {
    "资产编号":     "asset_tag",
    "品类":         "category",
    "工号":         "employee_id",
    "姓名":         "employee_name",
    "部门":         "department",
    "直属领导":     "supervisor",
    "资产名":       "hostname",
    "状态":         "status",
    "型号":         "model",
    "序列号":       "serial_number",
    "品牌":         "brand",
    "MAC地址":      "mac_address",
    "IP地址":       "ip_address",
    "备注":         "notes",
    "系统版本":     "system_version",
    "杀毒软件":     "antivirus_software",
    "锁号":         "lock_number",
    "位置":         "location",
    "数量":         "quantity",
    "采购日期":     "purchase_date",
    "固定资产编号": "fixed_asset_number",
    "PO号":         "po_number",
}

# 模糊列头别名表（常见变体 → 标准中文列头）
FUZZY_ALIAS: dict[str, str] = {
    # 资产编号
    "资产号": "资产编号", "编号": "资产编号", "资产标签": "资产编号",
    "asset_tag": "资产编号", "资产tag": "资产编号",
    # 品类
    "类别": "品类", "分类": "品类", "设备类型": "品类",
    "类型": "品类", "category": "品类",
    # 品牌
    "厂商": "品牌", "厂家": "品牌", "brand": "品牌",
    # 型号
    "规格": "型号", "产品型号": "型号", "model": "型号",
    "计算机型号": "型号", "机型": "型号",
    # 序列号
    "SN": "序列号", "sn": "序列号", "S/N": "序列号",
    "serial": "序列号", "序号": "序列号", "机器序列号": "序列号",
    # 状态
    "使用状态": "状态", "资产状态": "状态", "status": "状态",
    # 使用人（标准模板列头为“姓名”，旧列头继续作为别名）
    "使用人": "姓名", "员工姓名": "姓名", "姓名": "姓名", "用户": "姓名",
    "领用人": "姓名", "持有人": "姓名", "登记姓名": "姓名",
    # 直属领导
    "部门主管": "直属领导",
    # 工号
    "员工工号": "工号", "员工编号": "工号", "人员编号": "工号",
    # 部门
    "所属部门": "部门", "归属部门": "部门", "department": "部门",
    # 资产名/主机名
    "主机名": "资产名", "设备名": "资产名", "设备名称": "资产名",
    "hostname": "资产名", "计算机名": "资产名",
    # MAC 地址
    "mac": "MAC地址", "MAC": "MAC地址", "网卡地址": "MAC地址",
    "物理地址": "MAC地址", "Wireless mac": "MAC地址",
    # IP 地址
    "IP": "IP地址", "ip": "IP地址", "内网IP": "IP地址",
    # 固定资产编号
    "固资编号": "固定资产编号", "固定资产号": "固定资产编号",
    "财务编号": "固定资产编号", "Fixed assets": "资产编号",
    # 备注
    "说明": "备注", "描述": "备注", "remark": "备注", "notes": "备注",
    # 系统版本
    "操作系统": "系统版本", "OS": "系统版本",
    "os版本": "系统版本", "System": "系统版本",
    # 位置
    "存放位置": "位置", "资产位置": "位置", "放置位置": "位置",
    "使用位置": "位置", "FA使用区域": "位置",
    # 采购日期
    "购买日期": "采购日期", "入库日期": "采购日期", "购置日期": "采购日期",
    # PO 号
    "po号": "PO号", "PO": "PO号", "po": "PO号",
    "采购订单号": "PO号", "采购单号": "PO号", "订单号": "PO号",
}

# 必填列头（模糊匹配后仍需存在这两列）
REQUIRED_COLUMNS: frozenset[str] = frozenset({"资产编号", "状态"})

# 品类关键词推断规则（文件名 → 品类），按优先级排列
_FILENAME_CATEGORY_RULES: list[tuple[list[str], str]] = [
    (["笔记本", "laptop", "notebook", "nb"],         "笔记本电脑"),
    (["台式", "台式电脑", "desktop", "pc"],           "台式机"),
    (["服务器", "server", "srv"],                     "服务器"),
    (["移动设备", "pad", "ipad", "平板", "tablet"],   "平板电脑"),
    (["手机", "phone", "mobile"],                     "手机"),
    (["显示器", "monitor", "display"],                "显示器"),
    (["打印机", "printer"],                           "打印机"),
    (["网络", "network", "路由", "交换机", "switch"],  "网络设备"),
    (["鼠标", "mouse", "键盘", "keyboard"],           "无线鼠标"),
]

# 导入模板示例行
_TEMPLATE_EXAMPLE_ROW: dict[str, str] = {
    "资产编号":     "ZS-NB26-000001",
    "品类":         "笔记本电脑",
    "工号":         "",
    "姓名":         "",
    "部门":         "",
    "直属领导":     "",
    "资产名":       "LAPTOP-001",
    "状态":         "闲置",
    "型号":         "X1 Carbon",
    "序列号":       "SN123456789",
    "品牌":         "ThinkPad",
    "MAC地址":      "AA:BB:CC:DD:EE:FF",
    "IP地址":       "192.168.1.100",
    "备注":         "示例数据，请删除后填写实际数据",
    "系统版本":     "",
    "杀毒软件":     "",
    "锁号":         "",
    "位置":         "",
    "数量":         "",
    "采购日期":     "",
    "固定资产编号": "FA-2024-0001",
    "PO号":         "12000327",
}


# ===========================================================================
# ExcelSource
# ===========================================================================

class ExcelSource:
    """
    从 .xlsx 二进制内容读取资产数据，输出 AssetRecord 列表。

    使用方式：
        source = ExcelSource()
        records, inferred_category = source.read(file_bytes, filename)
    """

    def read(
        self,
        file_content: bytes,
        filename: str = "",
    ) -> tuple[list[AssetRecord], Optional[str]]:
        """
        解析 Excel 文件，构建 AssetRecord 列表。

        返回：
            (records, inferred_category)
            - records: AssetRecord 列表，每条记录 raw_fields 已填充（只读），
                       fields 填充英文字段名 → 原始字符串值
            - inferred_category: 从文件名推断的品类，无法推断时为 None

        异常：
            ValueError: 文件解析失败或缺少必填列头
        """
        # ── 1. 读取 Excel ──
        try:
            df = pd.read_excel(io.BytesIO(file_content), dtype=str)
        except Exception as exc:
            raise ValueError(f"文件解析失败: {exc}") from exc

        original_headers = list(df.columns)

        # ── 2. 必填列头检查 ──
        self._check_required_headers(original_headers)

        # ── 3. 文件名品类推断 ──
        inferred_category = self._infer_category_from_filename(filename)

        # ── 4. 构建列头解析映射 ──
        known_map, unknown_headers = self._build_header_map(original_headers)

        # ── 5. 构建 AssetRecord 列表 ──
        records: list[AssetRecord] = []
        for row_idx, (_, row) in enumerate(df.iterrows()):
            record = self._build_record(
                row=row,
                row_number=row_idx + 2,     # Excel 行号从 2 开始（1 为列头）
                filename=filename,
                known_map=known_map,
                unknown_headers=unknown_headers,
            )
            records.append(record)

        return records, inferred_category

    # ── 内部方法 ──────────────────────────────────────────────────────────

    def _check_required_headers(self, original_headers: list[str]) -> None:
        """检查必填列头是否存在（支持别名匹配）"""
        resolved_std_names: set[str] = set()
        for h in original_headers:
            h_stripped = str(h).strip()
            if h_stripped in COLUMN_MAPPING:
                resolved_std_names.add(h_stripped)
            elif h_stripped in FUZZY_ALIAS:
                resolved_std_names.add(FUZZY_ALIAS[h_stripped])

        missing = REQUIRED_COLUMNS - resolved_std_names
        if missing:
            raise ValueError(
                f"缺少必填列头: {', '.join(sorted(missing))}"
                f"（支持别名，如'编号'可代替'资产编号'）"
            )

    def _infer_category_from_filename(self, filename: str) -> Optional[str]:
        """根据文件名关键词推断品类，不区分大小写，按规则优先级匹配"""
        name_lower = filename.lower()
        for keywords, category in _FILENAME_CATEGORY_RULES:
            if any(kw.lower() in name_lower for kw in keywords):
                return category
        return None

    def _resolve_header(self, raw_header: str) -> Optional[str]:
        """
        将原始列头解析为英文字段名。

        解析顺序：
          1. 精确匹配 COLUMN_MAPPING
          2. 精确匹配 FUZZY_ALIAS → COLUMN_MAPPING
          3. 去除空格后重试 1、2
          4. 均未命中 → None（未知列）
        """
        h = str(raw_header).strip()

        if h in COLUMN_MAPPING:
            return COLUMN_MAPPING[h]
        if h in FUZZY_ALIAS:
            return COLUMN_MAPPING.get(FUZZY_ALIAS[h])

        h_nospace = h.replace(" ", "").replace("\u3000", "")
        if h_nospace in COLUMN_MAPPING:
            return COLUMN_MAPPING[h_nospace]
        if h_nospace in FUZZY_ALIAS:
            return COLUMN_MAPPING.get(FUZZY_ALIAS[h_nospace])

        return None

    def _build_header_map(
        self,
        original_headers: list[str],
    ) -> tuple[dict[str, str], list[str]]:
        """
        构建列头映射：
          known_map      : 原始列头 → 英文字段名（已知列，取首次命中）
          unknown_headers: 未知列的原始列头列表（存入 extra_fields）
        """
        known_map: dict[str, str] = {}
        unknown_headers: list[str] = []
        seen_en_fields: set[str] = set()

        for h in original_headers:
            en_field = self._resolve_header(str(h))
            if en_field is not None and en_field not in seen_en_fields:
                known_map[str(h)] = en_field
                seen_en_fields.add(en_field)
            else:
                unknown_headers.append(str(h))

        return known_map, unknown_headers

    def _read_cell(self, row: Any, header: str) -> Optional[str]:
        """读取单元格值，空值统一返回 None，其他保留原始字符串"""
        val = row[header]
        if not isinstance(val, str) and pd.isna(val):
            return None
        s = str(val).strip()
        return s if s else None

    def _build_record(
        self,
        row: Any,
        row_number: int,
        filename: str,
        known_map: dict[str, str],
        unknown_headers: list[str],
    ) -> AssetRecord:
        """
        从 DataFrame 行构建 AssetRecord。

        raw_fields : 原始列头 → 原始值（只读快照）
        fields     : 英文字段名 → 原始字符串值（后续 Normalizer 修改）
        extra_fields: 未知列 → 原始值
        """
        # raw_fields：原始快照，Key 用原始列头
        raw_fields: dict[str, Any] = {}
        for orig_h in list(known_map.keys()) + unknown_headers:
            raw_fields[orig_h] = self._read_cell(row, orig_h)

        # fields：英文字段名
        fields: dict[str, Any] = {}
        for orig_h, en_field in known_map.items():
            fields[en_field] = self._read_cell(row, orig_h)

        # extra_fields：未知列
        extra_fields: dict[str, Any] = {}
        for orig_h in unknown_headers:
            val = self._read_cell(row, orig_h)
            if val is not None:
                extra_fields[orig_h] = val

        record = AssetRecord(
            row_number=row_number,
            source_filename=filename,
        )
        # raw_fields 赋值后不再修改（约定，无语言级强制）
        object.__setattr__(record, "raw_fields", raw_fields)
        record.fields = fields
        record.extra_fields = extra_fields
        return record


# ===========================================================================
# 模板生成（独立函数，与 Source 读取逻辑分离）
# ===========================================================================

def generate_import_template() -> bytes:
    """
    生成导入模板 .xlsx 文件的二进制内容。

    模板包含：
      - 第 1 行：中文列头，加粗蓝色背景
      - 第 2 行：示例数据

    返回：
        .xlsx 文件的 bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "资产导入模板"

    headers = list(COLUMN_MAPPING.keys())
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"
    )
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.cell(row=2, column=col_idx, value=_TEMPLATE_EXAMPLE_ROW.get(header, ""))

        example_val = str(_TEMPLATE_EXAMPLE_ROW.get(header, ""))
        col_width = max(len(header) * 2, len(example_val) + 2, 12)
        ws.column_dimensions[cell.column_letter].width = col_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
