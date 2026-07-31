import io

from openpyxl import Workbook, load_workbook

import import_service
from import_v2.sources.excel_source import (
    COLUMN_MAPPING as V2_COLUMN_MAPPING,
    ExcelSource,
    generate_import_template,
)


EXPECTED_HEADERS = [
    "资产编号",
    "品类",
    "工号",
    "姓名",
    "部门",
    "直属领导",
    "资产名",
    "状态",
    "型号",
    "序列号",
    "品牌",
    "MAC地址",
    "IP地址",
    "备注",
    "系统版本",
    "杀毒软件",
    "锁号",
    "位置",
    "数量",
    "采购日期",
    "固定资产编号",
    "PO号",
]


def read_template_rows(content):
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))
    examples = list(next(rows))
    workbook.close()
    return headers, examples


def make_xlsx(headers, row):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(headers)
    worksheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_both_import_templates_use_requested_header_order_and_aligned_examples():
    for content in (generate_import_template(), import_service.generate_template()):
        headers, examples = read_template_rows(content)
        assert headers == EXPECTED_HEADERS
        values_by_header = dict(zip(headers, examples))
        assert values_by_header["资产编号"] == "ZS-NB26-000001"
        assert values_by_header["品类"] == "笔记本电脑"
        assert values_by_header["状态"] == "闲置"
        assert values_by_header["型号"] == "X1 Carbon"
        assert values_by_header["序列号"] == "SN123456789"
        assert values_by_header["品牌"] == "ThinkPad"
        assert values_by_header["备注"] == "示例数据，请删除后填写实际数据"
        assert values_by_header["固定资产编号"] == "FA-2024-0001"
        assert values_by_header["PO号"] == "12000327"

    assert list(V2_COLUMN_MAPPING) == EXPECTED_HEADERS
    assert list(import_service.COLUMN_MAPPING) == EXPECTED_HEADERS


def test_name_header_maps_to_employee_name_in_both_parsers():
    content = make_xlsx(
        ["资产编号", "品类", "姓名", "状态"],
        ["ZS-MR26-000901", "显示器", "张三", "闲置"],
    )

    records, _ = ExcelSource().read(content, "显示器清单.xlsx")
    assert records[0].fields["employee_name"] == "张三"

    _, rows, _ = import_service.parse_excel(content, "显示器清单.xlsx")
    assert rows[0]["employee_name"] == "张三"


def test_legacy_employee_name_headers_remain_supported():
    for legacy_header in ("使用人", "员工姓名"):
        content = make_xlsx(
            ["资产编号", "品类", legacy_header, "状态"],
            ["ZS-MR26-000902", "显示器", "李四", "闲置"],
        )

        records, _ = ExcelSource().read(content, "显示器清单.xlsx")
        assert records[0].fields["employee_name"] == "李四"

        _, rows, _ = import_service.parse_excel(content, "显示器清单.xlsx")
        assert rows[0]["employee_name"] == "李四"
